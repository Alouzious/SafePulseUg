import io
from datetime   import datetime
from openpyxl   import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment,
    Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# COLORS
# ─────────────────────────────────────────────────────────────
DARK_BLUE   = '1A3A5C'
MEDIUM_BLUE = '2563EB'
LIGHT_BLUE  = 'DBEAFE'
RED         = 'DC2626'
ORANGE      = 'F97316'
GREEN       = '16A34A'
GREY        = '6B7280'
LIGHT_GREY  = 'F3F4F6'
WHITE       = 'FFFFFF'


# ─────────────────────────────────────────────────────────────
# HELPER — Apply header style to a cell
# ─────────────────────────────────────────────────────────────
def style_header_cell(cell, bg_color=DARK_BLUE, font_color=WHITE, bold=True, size=10):
    cell.font       = Font(color=font_color, bold=bold, size=size, name='Calibri')
    cell.fill       = PatternFill(fill_type='solid', fgColor=bg_color)
    cell.alignment  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border     = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )


# ─────────────────────────────────────────────────────────────
# HELPER — Apply data cell style
# ─────────────────────────────────────────────────────────────
def style_data_cell(cell, bg_color=WHITE, bold=False, center=False):
    cell.font       = Font(size=9, name='Calibri', bold=bold)
    cell.fill       = PatternFill(fill_type='solid', fgColor=bg_color)
    cell.alignment  = Alignment(
        horizontal='center' if center else 'left',
        vertical='center',
        wrap_text=True
    )
    cell.border     = Border(
        left=Side(style='thin',   color='E5E7EB'),
        right=Side(style='thin',  color='E5E7EB'),
        top=Side(style='thin',    color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )


# ─────────────────────────────────────────────────────────────
# GENERATE CRIME LIST EXCEL
# ─────────────────────────────────────────────────────────────
def generate_crime_list_excel(reports, filters=None):
    """
    Generate an Excel report for a list of crime reports.
    Returns a BytesIO buffer.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Crime Reports"

    # ── Title rows ───────────────────────────────────────────
    ws.merge_cells('A1:J1')
    ws['A1'] = 'SAFEPULSE UG — UGANDA POLICE FORCE'
    ws['A1'].font       = Font(bold=True, size=16, color=WHITE, name='Calibri')
    ws['A1'].fill       = PatternFill(fill_type='solid', fgColor=DARK_BLUE)
    ws['A1'].alignment  = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:J2')
    ws['A2'] = f'Crime Reports — Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font       = Font(size=10, color=GREY, name='Calibri')
    ws['A2'].alignment  = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells('A3:J3')
    ws['A3'] = f'Total Records: {len(reports)} | Filters: {filters or "None"}'
    ws['A3'].font       = Font(size=9, color=GREY, name='Calibri', italic=True)
    ws['A3'].alignment  = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 16

    # ── Column headers ────────────────────────────────────────
    headers = [
        'Case Number', 'Title', 'Category', 'Severity',
        'Status', 'District', 'Location',
        'Date Occurred', 'Victim Count', 'Reported By'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        style_header_cell(cell)

    ws.row_dimensions[5].height = 20

    # ── Data rows ────────────────────────────────────────────
    severity_colors = {
        'low':      GREEN,
        'medium':   ORANGE,
        'high':     RED,
        'critical': '7C3AED',
    }

    for row_idx, r in enumerate(reports, 6):
        bg = LIGHT_GREY if row_idx % 2 == 0 else WHITE
        row_data = [
            r.case_number,
            r.title,
            r.category.replace('_', ' ').title(),
            r.severity.upper(),
            r.status.replace('_', ' ').title(),
            r.district,
            r.location,
            r.date_occurred.strftime('%Y-%m-%d %H:%M'),
            r.victim_count,
            r.reported_by.full_name if r.reported_by else 'Unknown',
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Color-code severity column
            if col_idx == 4:
                sev_color = severity_colors.get(r.severity, GREY)
                cell.font = Font(
                    bold=True, size=9,
                    color=WHITE, name='Calibri'
                )
                cell.fill = PatternFill(fill_type='solid', fgColor=sev_color)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color='E5E7EB'),
                    right=Side(style='thin', color='E5E7EB'),
                    top=Side(style='thin', color='E5E7EB'),
                    bottom=Side(style='thin', color='E5E7EB'),
                )
            else:
                style_data_cell(cell, bg_color=bg)

        ws.row_dimensions[row_idx].height = 18

    # ── Column widths ─────────────────────────────────────────
    col_widths = [18, 35, 18, 12, 22, 16, 30, 20, 14, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Stats sheet ──────────────────────────────────────────
    ws_stats = wb.create_sheet(title="Statistics")

    ws_stats.merge_cells('A1:D1')
    ws_stats['A1'] = 'CRIME STATISTICS SUMMARY'
    ws_stats['A1'].font      = Font(bold=True, size=14, color=WHITE, name='Calibri')
    ws_stats['A1'].fill      = PatternFill(fill_type='solid', fgColor=DARK_BLUE)
    ws_stats['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_stats.row_dimensions[1].height = 28

    # Count by category
    from collections import Counter
    cat_counts  = Counter(r.category for r in reports)
    stat_counts = Counter(r.status   for r in reports)
    sev_counts  = Counter(r.severity for r in reports)
    dis_counts  = Counter(r.district for r in reports)

    row = 3
    for section_title, counts in [
        ('BY CATEGORY', cat_counts),
        ('BY STATUS',   stat_counts),
        ('BY SEVERITY', sev_counts),
        ('BY DISTRICT', dis_counts),
    ]:
        ws_stats.merge_cells(f'A{row}:D{row}')
        ws_stats[f'A{row}'] = section_title
        style_header_cell(ws_stats[f'A{row}'], bg_color=MEDIUM_BLUE)
        ws_stats.row_dimensions[row].height = 18
        row += 1

        for label, count in counts.most_common():
            ws_stats.cell(row=row, column=1, value=label.replace('_', ' ').title())
            ws_stats.cell(row=row, column=2, value=count)
            style_data_cell(ws_stats.cell(row=row, column=1))
            style_data_cell(ws_stats.cell(row=row, column=2), center=True)
            row += 1
        row += 1

    for col in ['A', 'B', 'C', 'D']:
        ws_stats.column_dimensions[col].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer